import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl.reader.excel import load_workbook

book = 'add origin excel file here.xlsx'
book_loaded = load_workbook(book)
workbook = pd.ExcelFile(book)
df = workbook.parse(0, header=0)
dl = df['Domain'].values.tolist()
for i in range(len(dl[0])):
    # remove https if already in excel file
    url = 'https://' + dl[i]
    # some websites require headers for requests to view the website
    headers = {
        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 '
                      'Safari/537.36'}
    request = requests.get(url, headers=headers)
    soup = BeautifulSoup(request.text, 'html.parser')
    ti = soup.find('title').text
    hh = soup.findAll('h1')
    sheets = book_loaded.sheetnames
    sheet1 = book_loaded[sheets[0]]
    # column numbers should be changed to reflect data in your worksheet
    sheet1.cell(row=(i+2), column=3).value = ti
    for a in range(len(hh)):
        sheet1.cell(row=(i+2), column=4+a).value = hh[a].text
    # the found information will be written to the original file 
    book_loaded.save(book)



